# -*- coding: utf-8 -*-
"""
Liftlease — wekelijkse content-herinneringen
=============================================
Leest de Mediaplanning, zoekt de geplande posts voor de komende dagen,
en stuurt elke verantwoordelijke een mailtje met zijn/haar taken.

GEBRUIK
-------
  Testen (verstuurt NIETS, toont alleen wat er zou worden gemaild):
      python stuur_herinneringen.py --dry-run

  Echt versturen via Microsoft 365 (SMTP):
      python stuur_herinneringen.py --send

  KLAARZETTEN in je eigen Outlook (opent de mails; jij drukt zelf op Verzenden):
      python stuur_herinneringen.py --via outlook --klaarzetten

  Veilig TESTEN naar je eigen adres (zet alles klaar gericht aan jou):
      python stuur_herinneringen.py --via outlook --klaarzetten --test jij@liftlease.nl --datum 2026-01-05 --dagen 12

  Echt automatisch versturen (alleen als je dat expliciet wilt):
      python stuur_herinneringen.py --send --via outlook

  Andere peildatum / aantal dagen vooruit:
      python stuur_herinneringen.py --dry-run --datum 2026-01-01 --dagen 7

LET OP: voor 'echt versturen' moeten hieronder de e-mailadressen en de
afzender-gegevens (Microsoft 365) zijn ingevuld. Geen ervaring? Gebruik dan
liever de no-code variant via Power Automate (zie de handleiding).
"""

import argparse
import datetime as dt
import html
import json
import os
import re
import smtplib
import sys
import urllib.request
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import openpyxl

# =====================================================================
#  INSTELLINGEN  —  pas hier aan (dit is het enige wat je hoeft te wijzigen)
# =====================================================================

# Pad naar de planning. Standaard: het bestand naast dit script in de map erboven.
PLANNING_BESTAND = Path(__file__).resolve().parent.parent / "Mediaplanning 2026 - met recepten.xlsx"

# Waar staat de Content Maker-app? (komt in de mail als knop/link)
APP_URL = "https://VUL-HIER-DE-LINK-IN"   # de WEB-URL waar index.html draait (https). Leeg/placeholder = geen knop in de mail.
APP_URL_OK = APP_URL.startswith("http") and "VUL-HIER" not in APP_URL

# E-mailadres per medewerker (exact dezelfde naam als in kolom 'Verantwoordelijke')
MEDEWERKERS = {
    "Jan de Keijzer": "jan@liftlease.nl",
    "Niels":          "niels@liftlease.nl",
    "Gertjan":        "gertjan@liftlease.nl",
    "Liz":            "liz@liftlease.nl",
}

# Staat er geen naam in de kolom 'Verantwoordelijke'? Dan leiden we af uit het campagnetype
# (zelfde verdeling als in de Content Maker en de Recepten-legenda).
OWNER_BY_CAMPAGNE = {
    "Branding":      "Jan de Keijzer",
    "Afleveringen":  "Niels",
    "Commercieel":   "Niels",
    "Betrokkenheid": "Gertjan",
    "Ideeën":        "Liz",
}

# Naar wie gaan posts waar we écht niemand voor vinden? (bewaakt dat niets blijft liggen)
COORDINATOR = "liz@liftlease.nl"

# Afzender (Microsoft 365 / Outlook).
SMTP_SERVER   = "smtp.office365.com"
SMTP_POORT    = 587
AFZENDER      = "marketing@liftlease.nl"
AFZENDER_NAAM = "Liftlease Marketing"

# VEILIGHEID: zet het wachtwoord NIET in dit bestand. Het wordt uit een
# omgevingsvariabele gelezen. Stel die eenmalig in (PowerShell, eenmalig per pc):
#     setx LL_MAIL_WACHTWOORD "jouw-app-wachtwoord"
# Gebruik bij Microsoft 365 een app-wachtwoord. LET OP: veel M365-tenants hebben
# 'SMTP basic auth' uitgezet — werkt verzenden niet, gebruik dan Power Automate
# (zie handleiding) of Microsoft Graph i.p.v. dit script.
AFZENDER_WACHTWOORD = os.environ.get("LL_MAIL_WACHTWOORD", "")

# --- Lokale AI (LM Studio) ---------------------------------------------------
# Laat een lokaal model één persoonlijk openingszinnetje per mail schrijven.
# Zet in LM Studio: Developer -> model laden -> Start Server. Standaard draait die
# op http://localhost:1234. Staat de server UIT, dan valt het script automatisch
# terug op de standaardtekst (geen fout). Uitzetten kan ook met de vlag --geen-ai.
GEBRUIK_LOKALE_AI = True
# Gebruik 127.0.0.1 (precies zoals LM Studio toont bij 'Reachable at'); 'localhost' kan op
# Windows naar IPv6 (::1) wijzen en dan de verbinding weigeren.
AI_ENDPOINT = "http://127.0.0.1:1234/v1/chat/completions"
AI_MODEL = "google/gemma-4-e4b"   # de 'API Model Identifier' uit LM Studio (leeg = standaard)
AI_LAATSTE_FOUT = None            # interne diagnose: laatste foutmelding van de AI-aanroep

# =====================================================================
#  Vanaf hier hoef je niets te wijzigen
# =====================================================================

KOL = {"datum": 3, "campagne": 5, "kanaal": 6, "content": 7, "doel": 8,
       "verantwoordelijke": 11, "recept": 13}


def parse_datum(waarde):
    """Geeft een date terug uit een cel (datetime of tekst dd-mm-jjjj)."""
    if waarde is None:
        return None
    if isinstance(waarde, dt.datetime):
        return waarde.date()
    if isinstance(waarde, dt.date):
        return waarde
    s = str(waarde).strip()
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def lees_taken(pad, vandaag, dagen):
    """Verzamel posts met een datum in [vandaag, vandaag+dagen], gegroepeerd per persoon."""
    eind = vandaag + dt.timedelta(days=dagen)
    wb = openpyxl.load_workbook(pad, data_only=True)
    per_persoon = {}
    for ws in wb.worksheets:
        if ws.title not in ("Q1", "Q2", "Q3", "Q4"):
            continue
        for r in range(4, ws.max_row + 1):
            content = ws.cell(r, KOL["content"]).value
            if not content:
                continue  # lege agenderegel
            datum = parse_datum(ws.cell(r, KOL["datum"]).value)
            if datum is None or not (vandaag <= datum <= eind):
                continue
            naam = ws.cell(r, KOL["verantwoordelijke"]).value
            naam = str(naam).strip() if naam else ""
            if not naam:  # geen naam ingevuld -> afleiden uit campagnetype
                camp = ws.cell(r, KOL["campagne"]).value
                naam = OWNER_BY_CAMPAGNE.get(str(camp).strip(), "") if camp else ""
            if not naam:
                naam = "(niet toegewezen)"
            taak = {
                "datum": datum,
                "kanaal": (ws.cell(r, KOL["kanaal"]).value or "").strip() if isinstance(ws.cell(r, KOL["kanaal"]).value, str) else (ws.cell(r, KOL["kanaal"]).value or ""),
                "content": str(content).strip(),
                "recept": ws.cell(r, KOL["recept"]).value or "",
                "doel": ws.cell(r, KOL["doel"]).value or "",
            }
            per_persoon.setdefault(naam, []).append(taak)
    for taken in per_persoon.values():
        taken.sort(key=lambda t: t["datum"])
    return per_persoon


def _strip_aanhef(zin, voornaam):
    """Knipt een eventuele aanhef ('Hoi Niels,' / 'Hallo!') of losse naam vooraan weg en zet de
    eerste letter weer als hoofdletter. Voorkomt een dubbele aanhef in de mail."""
    zin = re.sub(r'^(hoi|hallo|hey|hi|beste|dag|goedemorgen|goedemiddag)\b[^,.!?\n]*[,!]?\s*', '', zin, flags=re.IGNORECASE)
    zin = re.sub(r'^' + re.escape(voornaam) + r'\s*[,!]\s*', '', zin, flags=re.IGNORECASE)
    zin = zin.strip().strip('"').strip()
    return (zin[0].upper() + zin[1:]) if zin else zin


def ai_zin(voornaam, taken):
    """Vraagt het lokale model (LM Studio) om één persoonlijke openingszin.
    Faalt stil (server uit / fout) -> geeft None terug, dan gebruikt de mail de standaardtekst."""
    global AI_LAATSTE_FOUT
    if not GEBRUIK_LOKALE_AI:
        return None
    taaklijst = "; ".join(f"{t['datum'].strftime('%d-%m')} {t['content']}" for t in taken)
    prompt = (
        "Je schrijft namens de marketing van Liftlease (intern transport specialist). "
        "Toon: nuchter, vakkundig, persoonlijk, kort. "
        f"Schrijf één vriendelijke, motiverende openingszin (max 20 woorden, Nederlands) over de content die collega {voornaam} "
        f"deze week maakt: {taaklijst}. "
        f"Begin NIET met een aanhef of de naam (dus niet 'Hoi', 'Hallo' of '{voornaam}'). "
        "Geef ALLEEN die ene zin terug — geen opsomming, geen emoji, geen aanhalingstekens."
    )
    payload = {
        "model": AI_MODEL or "local-model",
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.7,
        "max_tokens": 60,
    }
    req = urllib.request.Request(
        AI_ENDPOINT, data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json"},
    )
    try:
        with urllib.request.urlopen(req, timeout=120) as r:
            res = json.loads(r.read().decode("utf-8"))
        inhoud = res["choices"][0]["message"]["content"]
        regels = [r.strip().strip('"').strip() for r in inhoud.splitlines() if r.strip()]
        if not regels:
            return None
        return _strip_aanhef(regels[0], voornaam) or None
    except Exception as e:
        AI_LAATSTE_FOUT = f"{type(e).__name__}: {e}"
        return None


def maak_mail(naam, taken, intro=None):
    """Bouwt onderwerp + platte tekst + HTML voor één persoon. 'intro' = optionele AI-openingszin."""
    aantal = len(taken)
    onderwerp = f"📅 Jouw content deze week ({aantal} {'post' if aantal == 1 else 'posts'}) — Liftlease"

    regels_txt, regels_html = [], []
    for t in taken:
        d = t["datum"].strftime("%a %d-%m")
        recept = f"  [{t['recept']}]" if t["recept"] else ""
        regels_txt.append(f"• {d} — {t['kanaal']}: {t['content']}{recept}")
        # HTML-escapen: planningstekst kan < & e.d. bevatten -> nooit ongezuiverd in HTML
        e_content = html.escape(str(t["content"]))
        e_kanaal = html.escape(str(t["kanaal"]))
        e_recept = html.escape(str(t["recept"]))
        regels_html.append(
            f"<tr><td style='padding:6px 10px;border-bottom:1px solid #eee;white-space:nowrap'><b>{d}</b></td>"
            f"<td style='padding:6px 10px;border-bottom:1px solid #eee'>{e_content}"
            f"<br><span style='color:#5a6b73;font-size:12px'>{e_kanaal} · {e_recept}</span></td></tr>"
        )

    voornaam = naam.split()[0] if naam and naam != "(niet toegewezen)" else "team"
    voornaam_html = html.escape(voornaam)
    standaard = "Dit staat er deze week voor jou op de planning:"
    opening_txt = intro if intro else standaard
    opening_html = html.escape(intro) if intro else standaard
    txt = (
        f"Hoi {voornaam},\n\n"
        f"{opening_txt}\n\n"
        + "\n".join(regels_txt)
        + (f"\n\nMaak je posts in de Content Maker: {APP_URL}" if APP_URL_OK else "")
        + "\n\nHulp nodig met toon of opbouw? Kijk in het Content Playbook. Vragen? Check even met Liz."
        + "\n\nGroet,\nLiftlease Marketing"
    )
    app_btn = (f'<p style="margin-top:18px"><a href="{APP_URL}" style="background:#F2A03D;'
               f'color:#1b2b33;text-decoration:none;font-weight:bold;padding:11px 18px;'
               f'border-radius:9px;display:inline-block">Open de Content Maker →</a></p>') if APP_URL_OK else ""
    body_html = f"""\
<div style="font-family:Arial,sans-serif;color:#1C3B4A;max-width:560px">
  <div style="background:#1C3B4A;color:#fff;padding:14px 18px;border-radius:10px 10px 0 0">
    <span style="font-size:18px;font-weight:bold">Jouw content deze week</span>
    <span style="float:right;color:#F2A03D;font-weight:bold;font-size:12px">intern transport specialist</span>
  </div>
  <div style="border:1px solid #e1e7ea;border-top:none;border-radius:0 0 10px 10px;padding:16px 18px">
    <p>Hoi {voornaam_html},</p>
    <p>{opening_html}</p>
    <table style="border-collapse:collapse;width:100%;font-size:14px">{''.join(regels_html)}</table>
    {app_btn}
    <p style="color:#5a6b73;font-size:12px">Hulp met toon of opbouw? Kijk in het Content Playbook. Vragen? Check even met Liz.</p>
  </div>
</div>"""
    return onderwerp, txt, body_html


def verstuur(naar, onderwerp, txt, body_html):
    msg = MIMEMultipart("alternative")
    msg["Subject"] = onderwerp
    msg["From"] = f"{AFZENDER_NAAM} <{AFZENDER}>"
    msg["To"] = naar
    msg.attach(MIMEText(txt, "plain", "utf-8"))
    msg.attach(MIMEText(body_html, "html", "utf-8"))
    with smtplib.SMTP(SMTP_SERVER, SMTP_POORT) as s:
        s.starttls()
        s.login(AFZENDER, AFZENDER_WACHTWOORD)
        s.send_message(msg)


def verstuur_outlook(naar, onderwerp, body_html):
    """Verstuurt via de Outlook die op deze pc is geopend/ingelogd (geen wachtwoord nodig).
    Vereist: klassieke Outlook (desktop) + 'pip install pywin32'. Werkt niet met de 'nieuwe Outlook'-app."""
    import win32com.client  # lazy import: alleen nodig bij --via outlook
    ol = win32com.client.Dispatch("Outlook.Application")
    mail = ol.CreateItem(0)  # 0 = olMailItem
    mail.To = naar
    mail.Subject = onderwerp
    mail.HTMLBody = body_html
    mail.Send()


def klaarzetten_outlook(naar, onderwerp, body_html):
    """Maakt de mail aan en OPENT 'm in Outlook (verzendt NIET). Jij drukt zelf op Verzenden."""
    import win32com.client  # lazy import: alleen nodig bij --via outlook
    ol = win32com.client.Dispatch("Outlook.Application")
    mail = ol.CreateItem(0)  # 0 = olMailItem
    mail.To = naar
    mail.Subject = onderwerp
    mail.HTMLBody = body_html
    mail.Display(False)  # opent het venster (niet-modaal); niets wordt verzonden


def main():
    ap = argparse.ArgumentParser(description="Stuur wekelijkse content-herinneringen.")
    ap.add_argument("--send", action="store_true", help="Echt versturen (anders alleen tonen).")
    ap.add_argument("--dry-run", action="store_true", help="Alleen tonen, niets versturen (standaard).")
    ap.add_argument("--datum", help="Peildatum JJJJ-MM-DD (standaard: vandaag).")
    ap.add_argument("--dagen", type=int, default=7, help="Aantal dagen vooruit (standaard 7).")
    ap.add_argument("--via", choices=["smtp", "outlook"], default="smtp",
                    help="Verzendmethode: 'smtp' (M365-login) of 'outlook' (je eigen Outlook op deze pc, geen wachtwoord).")
    ap.add_argument("--klaarzetten", action="store_true",
                    help="Maak de mails klaar en OPEN ze in Outlook (niets verzonden; jij drukt zelf op Verzenden). Vereist --via outlook.")
    ap.add_argument("--test", metavar="EMAIL",
                    help="Zet/stuur ALLE mails naar dit testadres i.p.v. de echte ontvangers (veilig testen).")
    ap.add_argument("--geen-ai", action="store_true",
                    help="Schakel de lokale AI-openingszin uit en gebruik de standaardtekst.")
    args = ap.parse_args()

    vandaag = dt.datetime.strptime(args.datum, "%Y-%m-%d").date() if args.datum else dt.date.today()
    if args.klaarzetten and args.via != "outlook":
        sys.exit("--klaarzetten werkt alleen met --via outlook (het opent de mail in jouw Outlook).")
    klaarzetten = args.klaarzetten
    echt_versturen = args.send and not args.dry_run and not klaarzetten

    if not PLANNING_BESTAND.exists():
        sys.exit(f"Planning niet gevonden: {PLANNING_BESTAND}")

    per_persoon = lees_taken(PLANNING_BESTAND, vandaag, args.dagen)
    print(f"Peildatum {vandaag} · komende {args.dagen} dagen · "
          f"{sum(len(v) for v in per_persoon.values())} posts voor {len(per_persoon)} perso(o)n(en)\n")

    if not per_persoon:
        print("Geen geplande posts in deze periode. Klaar.")
        return

    if echt_versturen and args.via == "smtp" and not AFZENDER_WACHTWOORD:
        sys.exit("Geen wachtwoord gevonden. Zet de omgevingsvariabele LL_MAIL_WACHTWOORD "
                 "(zie bovenaan dit script), of gebruik --via outlook, of --dry-run.")

    ai_aan = GEBRUIK_LOKALE_AI and not args.geen_ai
    ai_mislukt = 0
    for naam, taken in per_persoon.items():
        adres = args.test or MEDEWERKERS.get(naam, COORDINATOR if naam == "(niet toegewezen)" else None)
        voornaam = naam.split()[0] if naam and naam != "(niet toegewezen)" else "team"
        intro = ai_zin(voornaam, taken) if ai_aan else None
        if ai_aan and not intro:
            ai_mislukt += 1
        onderwerp, txt, body_html = maak_mail(naam, taken, intro=intro)
        test_label = "  (TEST-adres)" if args.test else ""
        print("=" * 64)
        print(f"VOOR: {naam}{test_label}  ->  <{adres or 'GEEN E-MAILADRES — vul aan in MEDEWERKERS'}>")
        print(f"ONDERWERP: {onderwerp}")
        print(txt)
        print()
        if not (klaarzetten or echt_versturen):
            continue
        if not adres:
            print("  ! Overgeslagen: geen e-mailadres.")
            continue
        try:
            if klaarzetten:
                klaarzetten_outlook(adres, onderwerp, body_html)
                print("  ✓ Klaargezet in Outlook (nog NIET verzonden).")
            elif args.via == "outlook":
                verstuur_outlook(adres, onderwerp, body_html)
                print("  ✓ Verstuurd.")
            else:
                verstuur(adres, onderwerp, txt, body_html)
                print("  ✓ Verstuurd.")
        except smtplib.SMTPAuthenticationError:
            sys.exit("  ! Inloggen bij de mailserver mislukt. Waarschijnlijk staat 'SMTP basic auth' "
                     "uit in jullie Microsoft 365. Gebruik dan --via outlook of Power Automate (zie handleiding).")
        except Exception as e:
            sys.exit(f"  ! Mislukt: {e}  (Klassieke Outlook nodig + 'pip install pywin32'.)")

    if ai_aan and ai_mislukt:
        detail = f" Laatste fout: {AI_LAATSTE_FOUT}." if AI_LAATSTE_FOUT else ""
        print(f"(Lokale AI niet bereikbaar voor {ai_mislukt} mail(s) — standaardtekst gebruikt.{detail} "
              f"Server op {AI_ENDPOINT}?)")
    print("=" * 64)
    if klaarzetten:
        print("De mails staan nu klaar/open in Outlook. Controleer ze en druk zélf op Verzenden.")
    elif not echt_versturen:
        print("DIT WAS EEN TEST (--dry-run). Er is niets klaargezet of verstuurd.")
        print("Klaarzetten in Outlook? Voeg toe:  --via outlook --klaarzetten")


if __name__ == "__main__":
    main()
